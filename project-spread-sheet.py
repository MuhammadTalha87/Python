import openpyxl as xl

workbook = xl.load_workbook("marks-sheet.xlsx")

sheet = workbook["Sheet1"]

#for each row in range

for row in range(2, sheet.max_row + 1): #It starts from row 2 
    physics = sheet.cell(row, 3).value  #sheet.cell ( ) represents row and column
    maths = sheet.cell(row, 4).value
    cs = sheet.cell(row, 5).value
    project = sheet.cell(row, 6).value
    programming = sheet.cell(row, 7).value
    business_analaysis = sheet.cell(row, 8).value

    print(physics, maths, cs, project, programming, business_analaysis)
    totalmarks = physics + maths + cs + project + programming + business_analaysis
    totalmarks_cell = sheet.cell(row, 9) 
    totalmarks_cell.value = totalmarks
    average_marks_cell = sheet.cell(row, 10)
    average_marks = round(totalmarks/6)
    average_marks_cell.value = average_marks

workbook.save("student_percentage.xlsx")
print("Document saved successfully")


